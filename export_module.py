"""
Export-Modul: PDF und XLSX-Ausgabe der Rationsauswertung
"""

from pathlib import Path
from datetime import datetime
from typing import Optional


# ---------------------------------------------------------------------------
# XLSX Export
# ---------------------------------------------------------------------------

def export_xlsx(pferd: dict, bedarf, ist, differenz,
                positionen: list, ziel_pfad: str):
    """Exportiert die Rationsauswertung als Excel-Datei."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl ist nicht installiert.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rationsauswertung"

    # Farben
    header_fill = PatternFill("solid", fgColor="2E4057")
    sub_fill    = PatternFill("solid", fgColor="4A90D9")
    warn_fill   = PatternFill("solid", fgColor="FF6B6B")
    ok_fill     = PatternFill("solid", fgColor="6BCB77")
    neutral_fill= PatternFill("solid", fgColor="F0F0F0")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font  = Font(bold=True, size=14, color="2E4057")

    def schreibe_zeile(row, werte, fill=None, font=None, bold=False):
        for col, wert in enumerate(werte, 1):
            cell = ws.cell(row=row, column=col, value=wert)
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            elif bold:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left")
        return row + 1

    row = 1

    # Titel
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="Pferde-Rationsauswertung").font = title_font
    row += 1

    ws.cell(row=row, column=1,
            value=f"Erstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    row += 2

    # Pferdeinfo
    row = schreibe_zeile(row,
        ["Pferd", pferd.get("name",""), "Gewicht", f"{pferd.get('gewicht_kg',0):.0f} kg",
         "Nutzung", pferd.get("nutzung","")],
        fill=sub_fill, font=header_font)
    row += 1

    # Rations-Positionen
    row = schreibe_zeile(row,
        ["Futtermittel", "Menge FM (kg)", "TS (kg)", "Energie (MJ)", "Rohprotein (g)", ""],
        fill=header_fill, font=header_font)

    for pos in positionen:
        tm = pos.menge_kg * (1 - pos.wassergehalt_pct / 100)
        row = schreibe_zeile(row, [
            pos.name,
            f"{pos.menge_kg:.2f}",
            f"{tm:.2f}",
            f"{tm * pos.energie_mj_me:.1f}",
            f"{tm * pos.rohprotein_pct * 10:.0f}",
            ""
        ])
    row += 1

    # Vergleichstabelle
    headers = ["Parameter", "Einheit", "Bedarf", "Ist", "Differenz", "Diff %", "Status"]
    row = schreibe_zeile(row, headers, fill=header_fill, font=header_font)

    def diff_fill(diff_val, positiv_gut=True):
        if abs(diff_val) < 0.01:
            return neutral_fill
        if positiv_gut:
            return ok_fill if diff_val >= 0 else warn_fill
        else:
            return warn_fill if diff_val > 0 else ok_fill

    def status(diff_val, positiv_gut=True):
        if abs(diff_val) < 0.5:
            return "✓ gedeckt"
        if positiv_gut:
            return "✓ über Bedarf" if diff_val > 0 else "⚠ Unterversorgung"
        return "⚠ über Limit" if diff_val > 0 else "✓ unter Limit"

    kennzahlen = [
        ("Trockenmasse",   "kg/Tag",  bedarf.trockenmasse_kg,    ist.trockenmasse_kg,  True),
        ("Energie",        "MJ/Tag",  bedarf.energie_mj,         ist.energie_mj,       True),
        ("Rohprotein",     "g/Tag",   bedarf.rp_g,               ist.rohprotein_g,     True),
        ("Lysin",          "g/Tag",   bedarf.lysin_g,            ist.lysin_g,          True),
        ("Calcium",        "g/Tag",   bedarf.calcium_g,          ist.calcium_g,        True),
        ("Phosphor",       "g/Tag",   bedarf.phosphor_g,         ist.phosphor_g,       True),
        ("Magnesium",      "g/Tag",   bedarf.magnesium_g,        ist.magnesium_g,      True),
        ("Natrium",        "g/Tag",   bedarf.natrium_g,          ist.natrium_g,        True),
        ("Kupfer",         "mg/Tag",  bedarf.kupfer_mg,          ist.kupfer_mg,        True),
        ("Zink",           "mg/Tag",  bedarf.zink_mg,            ist.zink_mg,          True),
        ("Mangan",         "mg/Tag",  bedarf.mangan_mg,          ist.mangan_mg,        True),
        ("Selen",          "mg/Tag",  bedarf.selen_mg,           ist.selen_mg,         True),
        ("Vitamin E",      "mg/Tag",  bedarf.vit_e_mg,           ist.vit_e_mg,         True),
    ]

    for name, einheit, bed_val, ist_val, pos_gut in kennzahlen:
        diff_val = ist_val - bed_val
        diff_pct = (diff_val / bed_val * 100) if bed_val else 0.0
        fill = diff_fill(diff_val, pos_gut)
        row = schreibe_zeile(row, [
            name, einheit,
            f"{bed_val:.2f}", f"{ist_val:.2f}",
            f"{diff_val:+.2f}", f"{diff_pct:+.0f} %",
            status(diff_val, pos_gut)
        ], fill=fill)

    # NSC-Limitation
    if bedarf.nsc_max_pct:
        row += 1
        diff_nsc = ist.nsc_pct_von_ts - bedarf.nsc_max_pct
        fill = warn_fill if diff_nsc > 0 else ok_fill
        diff_nsc_pct = (diff_nsc / bedarf.nsc_max_pct * 100) if bedarf.nsc_max_pct else 0.0
        row = schreibe_zeile(row, [
            "NSC (Stärke+Zucker)", "% TS",
            f"max. {bedarf.nsc_max_pct:.0f}", f"{ist.nsc_pct_von_ts:.1f}",
            f"{diff_nsc:+.1f}", f"{diff_nsc_pct:+.0f} %",
            "⚠ ÜBERSCHRITTEN" if diff_nsc > 0 else "✓ OK"
        ], fill=fill)

    # Spaltenbreiten
    for col, breite in enumerate([30, 12, 10, 10, 10, 10, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = breite

    wb.save(ziel_pfad)


# ---------------------------------------------------------------------------
# PDF Export
# ---------------------------------------------------------------------------

def export_pdf(pferd: dict, bedarf, ist, differenz,
               positionen: list, ziel_pfad: str):
    """Exportiert die Rationsauswertung als PDF."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle)
    except ImportError:
        raise RuntimeError("reportlab ist nicht installiert.")

    doc = SimpleDocTemplate(ziel_pfad, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    # Farben
    DUNKELBLAU = colors.HexColor("#2E4057")
    HELLBLAU   = colors.HexColor("#4A90D9")
    ROT        = colors.HexColor("#FF6B6B")
    GRUEN      = colors.HexColor("#6BCB77")
    GRAU       = colors.HexColor("#F0F0F0")

    style_title = ParagraphStyle("title", parent=styles["Title"],
                                 textColor=DUNKELBLAU, fontSize=16)
    style_h2    = ParagraphStyle("h2", parent=styles["Heading2"],
                                 textColor=DUNKELBLAU)
    style_normal= styles["Normal"]

    # Titel
    story.append(Paragraph("Pferde-Rationsauswertung", style_title))
    story.append(Paragraph(
        f"Erstellt am {datetime.now().strftime('%d.%m.%Y %H:%M')} | "
        f"Pferd: <b>{pferd.get('name','')}</b> | "
        f"Gewicht: <b>{pferd.get('gewicht_kg',0):.0f} kg</b>",
        style_normal))
    story.append(Spacer(1, 0.5*cm))

    # Rations-Positionen
    story.append(Paragraph("Futterration", style_h2))
    pos_data = [["Futtermittel", "Menge FM (kg)", "TS (kg)", "Energie (MJ)", "RP (g)"]]
    for pos in positionen:
        tm = pos.menge_kg * (1 - pos.wassergehalt_pct / 100)
        pos_data.append([
            pos.name,
            f"{pos.menge_kg:.2f}",
            f"{tm:.2f}",
            f"{tm * pos.energie_mj_me:.1f}",
            f"{tm * pos.rohprotein_pct * 10:.0f}",
        ])

    pos_table = Table(pos_data, colWidths=[8*cm, 3*cm, 3*cm, 3*cm, 3*cm])
    pos_table.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), HELLBLAU),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, GRAU]),
        ("GRID",        (0,0), (-1,-1), 0.5, colors.lightgrey),
        ("FONTSIZE",    (0,0), (-1,-1), 9),
        ("ALIGN",       (1,0), (-1,-1), "CENTER"),
    ]))
    story.append(pos_table)
    story.append(Spacer(1, 0.5*cm))

    # Vergleichstabelle
    story.append(Paragraph("Nährstoffvergleich: Bedarf vs. Ist", style_h2))

    vergl_data = [["Parameter", "Einheit", "Bedarf", "Ist", "Differenz", "Status"]]
    kennzahlen = [
        ("Trockenmasse",  "kg/Tag",  bedarf.trockenmasse_kg, ist.trockenmasse_kg),
        ("Energie",       "MJ/Tag",  bedarf.energie_mj,      ist.energie_mj),
        ("Rohprotein",    "g/Tag",   bedarf.rp_g,            ist.rohprotein_g),
        ("Lysin",         "g/Tag",   bedarf.lysin_g,         ist.lysin_g),
        ("Calcium",       "g/Tag",   bedarf.calcium_g,       ist.calcium_g),
        ("Phosphor",      "g/Tag",   bedarf.phosphor_g,      ist.phosphor_g),
        ("Magnesium",     "g/Tag",   bedarf.magnesium_g,     ist.magnesium_g),
        ("Natrium",       "g/Tag",   bedarf.natrium_g,       ist.natrium_g),
        ("Kupfer",        "mg/Tag",  bedarf.kupfer_mg,       ist.kupfer_mg),
        ("Zink",          "mg/Tag",  bedarf.zink_mg,         ist.zink_mg),
        ("Mangan",        "mg/Tag",  bedarf.mangan_mg,       ist.mangan_mg),
        ("Selen",         "mg/Tag",  bedarf.selen_mg,        ist.selen_mg),
        ("Vitamin E",     "mg/Tag",  bedarf.vit_e_mg,        ist.vit_e_mg),
    ]

    row_colors = []
    for i, (name, einheit, bed_val, ist_val) in enumerate(kennzahlen, 1):
        diff = ist_val - bed_val
        status = "✓ gedeckt" if diff >= 0 else "⚠ Mangel"
        vergl_data.append([
            name, einheit,
            f"{bed_val:.2f}", f"{ist_val:.2f}",
            f"{diff:+.2f}", status
        ])
        row_colors.append(GRUEN if diff >= 0 else ROT)

    vergl_table = Table(vergl_data, colWidths=[4.5*cm, 2.5*cm, 3*cm, 3*cm, 3*cm, 4*cm])
    ts = TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), DUNKELBLAU),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID",        (0,0), (-1,-1), 0.5, colors.lightgrey),
        ("FONTSIZE",    (0,0), (-1,-1), 8),
        ("ALIGN",       (1,0), (-1,-1), "CENTER"),
    ])
    for i, farbe in enumerate(row_colors, 1):
        ts.add("BACKGROUND", (5, i), (5, i), farbe)

    vergl_table.setStyle(ts)
    story.append(vergl_table)

    # NSC-Hinweis
    if bedarf.nsc_max_pct:
        story.append(Spacer(1, 0.3*cm))
        nsc_ist = ist.nsc_pct_von_ts
        farbe = "red" if nsc_ist > bedarf.nsc_max_pct else "green"
        story.append(Paragraph(
            f'<font color="{farbe}"><b>NSC-Gehalt der Ration: {nsc_ist:.1f}% der TS '
            f'(Limit: {bedarf.nsc_max_pct:.0f}%)</b></font>',
            style_normal))

    # Fußzeile
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph(
        "Berechnung nach GfE-Empfehlungen. Alle Angaben ohne Gewähr. "
        "Bitte mit Tierarzt oder Fütterungsberater absprechen.",
        ParagraphStyle("footer", parent=styles["Normal"],
                       fontSize=7, textColor=colors.grey)))

    doc.build(story)


# ---------------------------------------------------------------------------
# Erhebungsblatt  –  druckbares Leerformular (A4)
# ---------------------------------------------------------------------------

def export_erhebungsblatt_pdf(ziel_pfad: str):
    """
    Erstellt ein druckbares A4-Leerformular zur Ist-Status-Erhebung im Stall.
    Kein Pferd-Objekt nötig – alle Felder sind zum Handausfüllen vorgesehen.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm, mm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable, KeepTogether
        )
        from reportlab.pdfbase import pdfmetrics
    except ImportError:
        raise RuntimeError("reportlab ist nicht installiert.")

    DUNKELBLAU = colors.HexColor("#2E4057")
    MITTELBLAU = colors.HexColor("#4A90D9")
    HELLBLAU   = colors.HexColor("#D0E8FF")
    HELLGRAU   = colors.HexColor("#F5F5F0")
    RAHMEN     = colors.HexColor("#B0B8C1")
    WEISS      = colors.white

    styles = getSampleStyleSheet()

    def S(name, **kw):
        base = styles.get(name) or styles["Normal"]
        return ParagraphStyle(f"_EB_{name}_{id(kw)}", parent=base, **kw)

    s_titel   = S("Title",   fontSize=15, textColor=DUNKELBLAU, spaceAfter=1*mm)
    s_sub     = S("Normal",  fontSize=8,  textColor=colors.grey, spaceAfter=3*mm)
    s_sekt    = S("Normal",  fontSize=9,  textColor=WEISS,
                  fontName="Helvetica-Bold", spaceAfter=0, spaceBefore=0)
    s_body    = S("Normal",  fontSize=8,  leading=12)
    s_klein   = S("Normal",  fontSize=7,  textColor=colors.grey,  leading=10)
    s_footer  = S("Normal",  fontSize=7,  textColor=colors.grey)

    doc = SimpleDocTemplate(
        ziel_pfad, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=1.8*cm, bottomMargin=2*cm
    )

    # -----------------------------------------------------------------------
    # Helfer
    # -----------------------------------------------------------------------
    PAGE_W = A4[0] - 3.6*cm   # Nutzbarer Seitenbereich

    def sektion(text: str):
        """Farbiger Abschnittsheader."""
        t = Table([[Paragraph(f"  {text}", s_sekt)]],
                  colWidths=[PAGE_W], rowHeights=[6*mm])
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, -1), DUNKELBLAU),
            ("TOPPADDING",  (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ]))
        return t

    def linie(breite=PAGE_W, farbe=RAHMEN, dicke=0.4):
        return HRFlowable(width=breite, thickness=dicke,
                          color=farbe, spaceAfter=1*mm, spaceBefore=0)

    def checkbox_zeile(*optionen, vorsatz=""):
        """Gibt 'vorsatz  □ Opt1  □ Opt2  ...' als Paragraph zurück."""
        def cb(label):
            return f"<font size='9'>□</font> {label}"
        zeile = ("  " + vorsatz + "  " if vorsatz else "") + \
                "     ".join(cb(o) for o in optionen)
        return Paragraph(zeile, s_body)

    def eingabe_zeile(label: str, breite_linie=8*cm, einheit=""):
        """Beschriftung + Strich + optionale Einheit als Tabellen-Zeile."""
        linie_zelle = Table([[""]],
                            colWidths=[breite_linie], rowHeights=[5*mm])
        linie_zelle.setStyle(TableStyle([
            ("LINEBELOW", (0, 0), (0, 0), 0.5, RAHMEN),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        row = [[Paragraph(label, s_body), linie_zelle]]
        if einheit:
            row[0].append(Paragraph(einheit, s_klein))
        return row

    def feingitter_tabelle(spalten: list, zeilen_anzahl: int,
                           col_breiten: list = None):
        """Leere Tabelle mit beschrifteten Spaltenköpfen."""
        header = [Paragraph(s, S("Normal", fontSize=7, fontName="Helvetica-Bold"))
                  for s in spalten]
        data = [header] + [[""] * len(spalten) for _ in range(zeilen_anzahl)]
        cw = col_breiten or [PAGE_W / len(spalten)] * len(spalten)
        t = Table(data, colWidths=cw, rowHeights=[6*mm] + [8*mm] * zeilen_anzahl)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), HELLBLAU),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("GRID",          (0, 0), (-1, -1), 0.4, RAHMEN),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [WEISS, HELLGRAU]),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return t

    # -----------------------------------------------------------------------
    # Story aufbauen
    # -----------------------------------------------------------------------
    story = []

    # === KOPF ===
    kopf_links = [
        [Paragraph("🐴  Pferde-Fütterungs-Erhebungsblatt", s_titel)],
        [Paragraph(
            "Zur Aufnahme des Ist-Zustands im Stall  |  "
            f"Formular-Version: {datetime.now().strftime('%d.%m.%Y')}",
            s_sub)],
    ]
    kopf_rechts = [
        [Paragraph("Datum der Erhebung:", s_body)],
        [Table([[""]], colWidths=[4*cm], rowHeights=[5*mm],
               style=[("LINEBELOW", (0,0), (0,0), 0.5, RAHMEN),
                      ("BOTTOMPADDING", (0,0), (-1,-1), 0)])],
        [Paragraph("Erhoben von:", s_body)],
        [Table([[""]], colWidths=[4*cm], rowHeights=[5*mm],
               style=[("LINEBELOW", (0,0), (0,0), 0.5, RAHMEN),
                      ("BOTTOMPADDING", (0,0), (-1,-1), 0)])],
    ]
    kopf = Table(
        [[Table(kopf_links, colWidths=[PAGE_W - 5*cm]),
          Table(kopf_rechts, colWidths=[5*cm])]],
        colWidths=[PAGE_W - 5*cm, 5*cm]
    )
    kopf.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(kopf)
    story.append(linie(dicke=1.5, farbe=DUNKELBLAU))

    # =========================================================
    # 1) PFERD & BESITZER
    # =========================================================
    story.append(sektion("1  |  Pferd & Betrieb"))
    story.append(Spacer(1, 2*mm))

    block1 = Table([
        eingabe_zeile("Name des Pferdes:", 8*cm) +
        eingabe_zeile("Besitzer / Betrieb:", 7.2*cm),
        eingabe_zeile("Rasse / Typ:", 7*cm) +
        eingabe_zeile("Stallbezeichnung / Standort:", 7.2*cm),
    ], colWidths=[3.5*cm, 8*cm, 4*cm, 7.2*cm])
    block1.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(block1)

    block2 = Table([
        eingabe_zeile("Körpergewicht:", 3.5*cm, "kg  (□ gewogen  □ geschätzt)") +
        eingabe_zeile("Alter:", 2.5*cm, "Jahre") +
        [Paragraph("BCS (1–9):", s_body),
         Table([[""]], colWidths=[1.5*cm], rowHeights=[5*mm],
               style=[("LINEBELOW",(0,0),(0,0),0.5,RAHMEN),
                      ("BOTTOMPADDING",(0,0),(-1,-1),0)])],
    ], colWidths=[2.8*cm, 3.5*cm, 4*cm, 2.0*cm, 3.5*cm, 2.5*cm, 2.2*cm, 1.5*cm])
    block2.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(block2)

    story.append(Spacer(1, 1*mm))
    story.append(checkbox_zeile("Stute", "Hengst", "Wallach",
                                vorsatz="Geschlecht:"))
    story.append(Spacer(1, 1.5*mm))
    story.append(checkbox_zeile(
        "Haltung / Rente", "Freizeitreiten", "Leichte Arbeit (<1 h/Tag)",
        "Mittlere Arbeit (1–2 h/Tag)", "Schwere Arbeit (>2 h/Tag)",
        vorsatz="Nutzung:"))
    story.append(Spacer(1, 1.5*mm))

    tr_lak = Table([
        [Paragraph("Trächtigkeit:", s_body),
         checkbox_zeile("nein", "ja, Monat:"),
         Table([[""]], colWidths=[1.5*cm], rowHeights=[5*mm],
               style=[("LINEBELOW",(0,0),(0,0),0.5,RAHMEN),
                      ("BOTTOMPADDING",(0,0),(-1,-1),0)]),
         Paragraph("", s_body),
         Paragraph("Laktation:", s_body),
         checkbox_zeile("nein", "ja, Monat:"),
         Table([[""]], colWidths=[1.5*cm], rowHeights=[5*mm],
               style=[("LINEBELOW",(0,0),(0,0),0.5,RAHMEN),
                      ("BOTTOMPADDING",(0,0),(-1,-1),0)]),
         ],
    ], colWidths=[2.5*cm, 4.5*cm, 1.5*cm, 0.5*cm, 2.5*cm, 4*cm, 1.5*cm])
    tr_lak.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tr_lak)
    story.append(Spacer(1, 2*mm))

    # =========================================================
    # 2) GESUNDHEIT / DIAGNOSEN
    # =========================================================
    story.append(sektion("2  |  Gesundheit & besondere Diagnosen"))
    story.append(Spacer(1, 2*mm))
    story.append(checkbox_zeile(
        "EMS", "Cushing / PPID", "PSSM1", "PSSM2 / MIM",
        "Hufrehe (akut)", "Hufrehe (chron.)", "Magengeschwüre"))
    story.append(Spacer(1, 1.5*mm))
    story.append(checkbox_zeile(
        "Arthrose / Lahmheit", "Zahnprobleme / Schluckbeschwerden",
        "Atemwegserkrankung (COPD/RAO)", "Herzprobleme"))
    story.append(Spacer(1, 1.5*mm))

    diag_sonst = Table([
        eingabe_zeile("Sonstige Diagnosen / Medikamente:", PAGE_W - 4.5*cm)
    ], colWidths=[4.5*cm, PAGE_W - 4.5*cm])
    diag_sonst.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(diag_sonst)
    story.append(Spacer(1, 2*mm))

    # =========================================================
    # 3) HALTUNG & BEWEGUNG
    # =========================================================
    story.append(sektion("3  |  Haltung & Bewegung"))
    story.append(Spacer(1, 2*mm))
    story.append(checkbox_zeile(
        "Einzelbox", "Gruppenbox", "Offenstall", "Paddockbox",
        "Aktivstall / Laufstall",
        vorsatz="Haltungsform:"))
    story.append(Spacer(1, 1.5*mm))
    story.append(checkbox_zeile(
        "Selbsttränke (unbegrenzt)", "Eimer (Menge/Tag: ____ L)",
        "Bach / Teich / Naturquelle",
        vorsatz="Wasser:"))
    story.append(Spacer(1, 1.5*mm))

    bew = Table([
        [Paragraph("Bewegung/Arbeit pro Tag:", s_body),
         checkbox_zeile("keine", "< 30 min", "30–60 min", "1–2 h", "> 2 h"),
         Paragraph("Weidegang täglich:", s_body),
         Table([[""]], colWidths=[1.5*cm], rowHeights=[5*mm],
               style=[("LINEBELOW",(0,0),(0,0),0.5,RAHMEN),
                      ("BOTTOMPADDING",(0,0),(-1,-1),0)]),
         Paragraph("h", s_body)],
    ], colWidths=[4.5*cm, 6.5*cm, 4*cm, 1.5*cm, 0.5*cm])
    bew.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(bew)
    story.append(Spacer(1, 2*mm))

    # =========================================================
    # 4) WEIDE
    # =========================================================
    story.append(sektion("4  |  Weide & Grünfutter"))
    story.append(Spacer(1, 2*mm))
    story.append(checkbox_zeile(
        "kein Weidegang", "Koppel / Paddock (kein Bewuchs)",
        "Weide (begrenzter Bewuchs)", "Weide (üppiger Bewuchs)",
        vorsatz="Weidetyp:"))
    story.append(Spacer(1, 1.5*mm))
    story.append(checkbox_zeile(
        "ganzjährig",
        "Saison: ___________ bis ___________",
        vorsatz="Weideperiode:"))
    story.append(Spacer(1, 1.5*mm))
    story.append(checkbox_zeile(
        "morgens (<6 h)", "mittags (<6 h)", "abends (<6 h)", "ganztags",
        vorsatz="Weide-Zeiten:"))
    story.append(Spacer(1, 2*mm))

    # =========================================================
    # 5) GRUNDFUTTER (Heu / Raufutter)
    # =========================================================
    story.append(sektion("5  |  Grundfutter / Heu / Raufutter"))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        "Alle Raufutter & heufähige Produkte (Heu, Heulage, Stroh, Heucobs, "
        "Nassfutter, Silage usw.)",
        s_klein))
    story.append(Spacer(1, 1.5*mm))
    story.append(feingitter_tabelle(
        spalten=[
            "Produkt / Beschreibung",
            "Qualität\n(sehr gut / gut / mittel / gering)",
            "Menge\nkg / Tag (FM)",
            "Anmerkung",
        ],
        zeilen_anzahl=5,
        col_breiten=[6.5*cm, 5.5*cm, 2.5*cm, 4.5*cm],
    ))
    story.append(Spacer(1, 2*mm))

    # =========================================================
    # 6) KRAFTFUTTER & ERGÄNZUNGEN
    # =========================================================
    story.append(sektion("6  |  Kraftfutter, Ergänzungsfutter & Mineralien"))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        "Alle weiteren Futtermittel: Müsli, Pellets, Mineralfutter, Öle, "
        "Ergänzungen, Kräuter, Leckerlis usw.",
        s_klein))
    story.append(Spacer(1, 1.5*mm))
    story.append(feingitter_tabelle(
        spalten=[
            "Produkt / Handelsname",
            "Hersteller",
            "Menge\ng oder kg / Tag",
            "Verabreichung\n(morgens / abends / gesamt)",
            "Anmerkung",
        ],
        zeilen_anzahl=9,
        col_breiten=[5.5*cm, 3.5*cm, 2.8*cm, 4*cm, 3.2*cm],
    ))
    story.append(Spacer(1, 2*mm))

    # =========================================================
    # 7) BEOBACHTUNGEN & ZUSTAND
    # =========================================================
    story.append(sektion("7  |  Beobachtungen & Allgemeinzustand"))
    story.append(Spacer(1, 2*mm))

    beob_data = [
        [Paragraph("Fressverhalten:", s_body),
         checkbox_zeile("normal", "wählerisch", "hastig / gierig",
                        "träge / wenig Appetit", "Kotwasser"),
         ],
        [Paragraph("Kotbeschaffenheit:", s_body),
         checkbox_zeile("normal", "zu fest / trocken",
                        "zu weich / breiig", "Pfützenbildung", "unregelmäßig"),
         ],
        [Paragraph("Fellbild:", s_body),
         checkbox_zeile("glänzend / gepflegt", "matt / stumpf",
                        "Sommerfellwechsel verspätet", "auffällig"),
         ],
        [Paragraph("Körperkondition:", s_body),
         checkbox_zeile("sehr gut", "gut", "mittel / dürr",
                        "zu dünn (BCS ≤ 4)", "zu fett (BCS ≥ 7)"),
         ],
        [Paragraph("Allgemeinzustand:", s_body),
         checkbox_zeile("sehr gut", "gut", "mittel",
                        "schlecht / krank"),
         ],
    ]
    beob_table = Table(beob_data,
                       colWidths=[3.8*cm, PAGE_W - 3.8*cm])
    beob_table.setStyle(TableStyle([
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("ROWBACKGROUNDS",(0, 0), (-1, -1), [WEISS, HELLGRAU]),
        ("LINEBELOW",     (0, 0), (-1, -1), 0.3, RAHMEN),
    ]))
    story.append(beob_table)
    story.append(Spacer(1, 2*mm))

    # =========================================================
    # 8) NOTIZEN / FREITEXT
    # =========================================================
    story.append(sektion("8  |  Notizen & besondere Beobachtungen"))
    story.append(Spacer(1, 1.5*mm))

    # Schreib-Zeilen
    notiz_zeilen = 7
    notiz_data = [[""] for _ in range(notiz_zeilen)]
    notiz_table = Table(notiz_data,
                        colWidths=[PAGE_W],
                        rowHeights=[8*mm] * notiz_zeilen)
    notiz_table.setStyle(TableStyle([
        ("LINEBELOW",     (0, 0), (-1, -1), 0.5, RAHMEN),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
    ]))
    story.append(notiz_table)
    story.append(Spacer(1, 3*mm))

    # =========================================================
    # 9) UNTERSCHRIFT
    # =========================================================
    story.append(linie(dicke=0.8, farbe=DUNKELBLAU))
    unt_data = [[
        Table([
            [Paragraph("Ort, Datum:", s_body)],
            [Table([[""]], colWidths=[5.5*cm], rowHeights=[5*mm],
                   style=[("LINEBELOW",(0,0),(0,0),0.5,RAHMEN),
                          ("BOTTOMPADDING",(0,0),(-1,-1),0)])],
        ], colWidths=[7*cm]),
        Table([
            [Paragraph("Unterschrift / Kürzel der erhebenden Person:", s_body)],
            [Table([[""]], colWidths=[8*cm], rowHeights=[5*mm],
                   style=[("LINEBELOW",(0,0),(0,0),0.5,RAHMEN),
                          ("BOTTOMPADDING",(0,0),(-1,-1),0)])],
        ], colWidths=[PAGE_W - 7*cm]),
    ]]
    unt_table = Table(unt_data, colWidths=[7*cm, PAGE_W - 7*cm])
    unt_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(unt_table)
    story.append(Spacer(1, 2*mm))

    # Fußzeile
    story.append(linie(dicke=0.4))
    story.append(Paragraph(
        "Gurktaler Pferdefutter-Rationsrechner  |  "
        "Dieses Formular dient der internen Datenhaltung und Fütterungsberatung.  |  "
        "Alle Angaben vertraulich behandeln.",
        s_footer))

    doc.build(story)
